#######################################################################
# NCT Creation Main
#
# This acts as the entry point for the program and calls any other scripts as needed.
#######################################################################
OutputSaveLocation = "Created NCTs.xlsx"
OutputNCTColumn = 1
OutputVINColumn = 2
OutputSNColumn = 3
#######################################################################

import tkinter
from tkinter import messagebox
from NCTInterpreter import LoadAndInterpretGSARData
from NCTData import *
from openpyxl import Workbook
#from openpyxl import worksheet

def __formatBook(wb:Workbook):
    ws = wb.active
    ws.title = "Created NCTs"
    ws.cell(row=1, column=OutputNCTColumn).value = "NCT #"
    ws.cell(row=1, column=OutputVINColumn).value = "VIN"
    ws.cell(row=1, column=OutputSNColumn).value = "S/N"

    ws2 = wb.create_sheet("Not Created")
    ws2.cell(row=1, column=1).value = "VIN"
    ws2.cell(row=1, column=2).value = "S/N"
    ws2.cell(row=1, column=3).value = "Excel Sheet Row"
    return ws2

#This is where NCT data is interpreted and uploaded to NCT.
if messagebox.askyesno("NCT","Are you sure you want to upload to NCT?"):
    NCTData = LoadAndInterpretGSARData() #An array of NCTDataContainers that was parsed from the file.
    wb = Workbook()
    ws = wb.active
    wsNotCreated = __formatBook(wb)
    nextRow = 2
    nextNotCreatedRow = 2
    for data in NCTData:
        NCT = data.CreateNCT()
        if isinstance(NCT, int):
            #NCT was created!
            ws.cell(row=nextRow, column=OutputNCTColumn).value = NCT
            ws.cell(row=nextRow, column=OutputVINColumn).value = data.getData(NCTDataType.VIN)
            ws.cell(row=nextRow, column=OutputSNColumn).value = data.getData(NCTDataType.SerialNumber)
            wb.save(OutputSaveLocation)
            nextRow = nextRow + 1
        else:
            #NCT was not created, returned was a list of missing fields.
            wsNotCreated.cell(row=nextNotCreatedRow, column=1).value = data.getData(NCTDataType.VIN)
            wsNotCreated.cell(row=nextNotCreatedRow, column=2).value = data.getData(NCTDataType.SerialNumber)
            wsNotCreated.cell(row=nextNotCreatedRow, column=3).value = data.GetExcelSheetRow()
            wb.save(OutputSaveLocation)
            nextNotCreatedRow = nextNotCreatedRow + 1

wb.save(OutputSaveLocation)
messagebox.showinfo("Status", "Program finished, saved to " + OutputSaveLocation)