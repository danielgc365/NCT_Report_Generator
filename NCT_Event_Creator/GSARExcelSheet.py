#######################################################################
# GSAR Excel Sheet

# Contains definitions for where data is stored in the excel sheet and
# convenient functions to access them.
#######################################################################

from enum import Enum
from tkinter import filedialog
from openpyxl import load_workbook

#This enum stores the column that each data is located in. This allows easy swapping if something moves.
class GSARExcelDataLocation(Enum):
    VIN = 1
    CSN = 2
    VeoneerProductionDate = 3
    CustomerPN = 4
    VehicleLine = 5
    DTCs = 6
    HistoricDTCs = 7
    Chargeback = 8
    PreAnalysisResults = 9
    DateOfArrival = 10
    ModelYear = 11
    CustomerPickList = 12
    VehicleProductionDate = 13
    WSD = 14
    RejectDate = 15
    TIS_WSD = 16 #I don't know what this is.
    MileageMiles = 17
    OdometerKM = 18
    DealerState = 19
    Dealer = 20
    ClaimInfo = 21
    DealerCountry = 22
    CustomerFailureCode = 23
    ComplaintSummary = 24
    ProductFamily = 25
    VeoneerFacility = 26
    CustomerPlant = 27
    Summary = 28
    IssueCategory = 29
    Project = 30
    VeoneerOwner = 31
    AssignedUser = 32
    ProductArea = 33
    Category8D = 34

#When created, prompts the user to load an Excel sheet as input
class GSARExcelWorksheet:
    def __init__(self):
        #Load the excel file
        filepath = filedialog.askopenfilename(title="Select file")
        wb = load_workbook(filepath, read_only=True, data_only=True)
        #TODO check if wb does not load properly
        #Select the correct worksheet
        #sheetnames = wb.sheetnames
        # TODO Create "Select Sheet" window
        self.Worksheet = wb["Sheet1"]
        
        self.MaxRow = self.Worksheet.max_row
        self.MaxCol = self.Worksheet.max_column
    
    def GetValue(self, Row:int, dataType:GSARExcelDataLocation):
        return self.Worksheet.cell(row=Row, column=int(dataType.value)).value
