#######################################################################
# Read Excel file
# Create an NCT Item per line in said document.
#######################################################################

from openpyxl import load_workbook
import datetime
from subprocess import check_output
from tkinter import *
from tkinter import filedialog

# Global variables
MasterList = []
Chargebacks8D = ["[825772]: CHARGEBACK - EXT-DEVICE OPEN/SHORT",
                 "[825775]: CHARGEBACK - COMMUNICATION FAULT",
                 "[825777]: CHARGEBACK - CONFIGURATION ERROR",
                 "[825778]: CHARGEBACK - EDR_FULL_AND_LOADED",
                 "[825779]: CHARGEBACK - ENS_OPEN_SHORT_TO_GND",
                 "[825780]: CHARGEBACK - IGNITION LOW",
                 "[825781]: CHARGEBACK - KNOWN ISSUE",
                 "[825782]: CHARGEBACK - NTF",
                 "[825783]: CHARGEBACK - SERVICE MODULE",
                 "[825784]: CHARGEBACK - VIN NOT PROGRAMMED",
                 "[825785]: CHARGEBACK - WATER CONTAMINATION",
                 "[825786]: CHARGEBACK - MODULE DAMAGED",
                 "[825787]: CHARGEBACK - WRONG PART RETURNED",
                 "[833660]: CHARGEBACK - OCS FAULT",
                 "[837056]: CHARGEBACK - ABS",
                 "[843433]: CHARGEBACK - SYSTEMS",
                 "[852906]: CHARGEBACK - OUT OF WARRANTY",
                 "[864387]: CHARGEBACK - Bosch Gyro Channel Fault"
                 ]
NCT_OutputList = [["NCT", "VIN", "Serial Number"]]
Tk().withdraw()
filepath = filedialog.askopenfilename(title="Select file")
wb = load_workbook(filepath, read_only=True, data_only=True)
sheetnames = wb.sheetnames
# TODO Create "Select Sheet" window
ws = wb["Sheet1"]
max_row = ws.max_row
max_col = ws.max_column

# Populate Master List
for i in range(1, max_row+1):
    item_list = []
    countFlag = 0
    for j in range(1, max_col+1):
        cell = ws.cell(row=i, column=j)
        item_list.append(cell.value)
    for item in item_list:
        if item == "":
            countFlag += 1
    if countFlag > 5:
        break
    MasterList.append(item_list)

# Create query variables
for indexMaster, itemMaster in enumerate(MasterList):
    if indexMaster > 0:
        ToBePopped = []
        for index, item in enumerate(MasterList[indexMaster]):
            NCT_Specific = []
            if isinstance(item, datetime.datetime):
                item = item.strftime("%b %d, %Y")
            if item is None:
                ToBePopped.append(index)
            elif index == 0:  # VIN
                NCT_Specific.append(item)
                MasterList[indexMaster][index] = "--field=\'AI_Chassi number="+MasterList[indexMaster][index]+"\'"
            elif index == 1:  # Serial Number
                NCT_Specific.append(item)
                MasterList[indexMaster][index] = "--field=\'AI_CSN number="+MasterList[indexMaster][index]+"\'"
            elif index == 2:  # Production Date
                MasterList[indexMaster][index] = "--field=\'AI_Production date=" + item + "\'"
            elif index == 3:  # Customer PN
                MasterList[indexMaster][index] = "--field=\'AI_Customer Part No=" + item + "\'" +\
                                                 " --field=\'AI_Part No=" + item + "\'"
            elif index == 4:  # Vehicle Line
                MasterList[indexMaster][index] = "--field=\'FA_Vehicle Line (Code)=" + item.upper() + "\'"
            elif index == 5:  # Active DTC
                    MasterList[indexMaster][index] = "--field=\'AI_DTC="+MasterList[indexMaster][index]+"\'"
            elif index == 6:  # Historic DTC
                    MasterList[indexMaster][index] = "--field=\'FA_Historic DTCs="+MasterList[indexMaster][index]+"\'"
            elif index == 7:  # Ignored
                ToBePopped.append(index)
            elif index == 8:  # Pre-Analysis Results
                    MasterList[indexMaster][index] = "--field=\'Pre-Analysis Results=" + \
                                                     MasterList[indexMaster][index] + "\'"
            elif index == 9:  # Date of arrival
                    MasterList[indexMaster][index] = "--field=\'AI_Date of Arrival=" + item + "\'"
            elif index == 10:  # Model Year
                MasterList[indexMaster][index] = "--field=\'FA_MY (Model Year)="+str(MasterList[indexMaster][index])+"\'"
            elif index == 11:  # Customer
                MasterList[indexMaster][index] = "--field=\'AI_Customer (pick list)="+item.capitalize()+"\'"
            elif index == 12:  # Vehicle Production Date
                MasterList[indexMaster][index] = "--field=\'Vehicle production date=" + item + "\'"
            elif index == 13:  # Vehicle Registration Date
                MasterList[indexMaster][index] = "--field=\'Vehicle registration date=" + item + "\'"
            elif index == 14:  # Reject Date
                MasterList[indexMaster][index] = "--field=\'FA_Customer Reject Date=" + str(item) + "\'" +\
                                                 " --field=\'Error discovered date=" + str(item) + "\'" +\
                                                 " --field=\'FA_First 8D Sent Date=" + str(item) + "\'"
            elif index == 15:  # Ignored
                ToBePopped.append(index)
            elif index == 16:  # Ignored
                ToBePopped.append(index)
            elif index == 17:  # Odometer in Km
                MasterList[indexMaster][index] = "--field=\'AI_Km="+str(int(MasterList[indexMaster][index]))+"\'"
            elif index == 18:  # Dealer State
                MasterList[indexMaster][index] = "--field=\'FA_Dealer State [US Only]=" + \
                                                     MasterList[indexMaster][index]+"\'"
            elif index == 19:  # DealerName
                MasterList[indexMaster][index] = "--field=\'FA_Dealer="+MasterList[indexMaster][index]+"\'"
            elif index == 20:  # Claim Info
                MasterList[indexMaster][index] = "--field=\'AI_Claim info="+MasterList[indexMaster][index]+"\'"
            elif index == 21:  # Dealer Country
                if item in "CAN":
                    item = "Canada"
                MasterList[indexMaster][index] = "--field=\'FA_Dealer Country=" + item + "\'"
            elif index == 22:  # Customer Failure Code
                if isinstance(item, int):
                    MasterList[indexMaster][index] = "--field=\'FA_Customer Failure Code=" + \
                                                     str(MasterList[indexMaster][index])+"\'"
                else:
                    MasterList[indexMaster][index] = "--field=\'FA_Customer Failure Code=" + \
                                                     MasterList[indexMaster][index]+"\'"
            elif index == 23:  # Complaint Summary
                MasterList[indexMaster][index] = "--field=\'AI_Complaint Summary="+MasterList[indexMaster][index]+"\'"
            elif index == 24:  # Product Family
                MasterList[indexMaster][index] = "--field=\'FA_Product Family="+MasterList[indexMaster][index]+"\'"
            elif index == 25:  # Veoneer Facility
                if item.lower() in "cmm - veoneer canada markham":
                    MasterList[indexMaster][index] = "--field=\'AI_Country=" + "CMM - Veoneer Canada Markham" + "\'"
                elif item.lower() in "frm - veoneer france rouen":
                    MasterList[indexMaster][index] = "--field=\'AI_Country=" + "FRM - Veoneer France Rouen" + "\'"
                elif item.lower() in "cfm - veoneer china fengxian":
                    MasterList[indexMaster][index] = "--field=\'AI_Country=" + "CFM - Veoneer China Fengxian" + "\'"
            elif index == 26:  # Error Discovered
                MasterList[indexMaster][index] = "--field=\'AI_Error discovered (place)=" +\
                                                 MasterList[indexMaster][index]+"\'"
            elif index == 27:  # Summary
                MasterList[indexMaster][index] = "--field=\'Summary="+MasterList[indexMaster][index]+"\'"
            elif index == 28:  # Issue Category
                MasterList[indexMaster][index] = "--field=\'FA_Issue Category="+item.capitalize()+"\'"
            elif index == 29:  # Project
                item = "/Analysis/RCS Analysis"
                MasterList[indexMaster][index] = "--field=\'Project="+item+"\'"
            elif index == 30:  # Veoneer Owner
                MasterList[indexMaster][index] = "--field=\'CQE="+item[item.find("(") + 1:item.find(")")]+"\'"
            elif index == 31:  # Assigned User
                    MasterList[indexMaster][index] = "--field=\'Assigned User=" +item[item.find("(") + 1:item.find(")")]+ "\'"
            elif index == 32:  # Product Line and Area
                MasterList[indexMaster][index] = "--field=\'FA_Product Line=" + MasterList[indexMaster][index] + "\'" +\
                                                 " --field=\'AI_Product Area=" + item + "\'"
            elif index == 33:  # Chargeback or Under Analysis
                if MasterList[indexMaster][7].lower() in "y":
                    for FTF in Chargebacks8D:
                        if item in FTF:
                            item = FTF
                    MasterList[indexMaster][index] = "--field=\'AI_Fault Type Others=" + item + "\'" +\
                                                     " --field=\'AI_Customer Status=Closed\'" +\
                                                     " --field=\'AI_Current Location=CHARGEBACK\'"
                else:
                    item = "[840805]: ECU - UNDER ANALYSIS"
                    MasterList[indexMaster][index] = "--field=\'AI_Fault Type Others=" + item + "\'" +\
                                                     " --field=\'AI_Current Location=UST\'"
            elif index == 34:  # Ignored
                ToBePopped.append(index)
        for popItem in sorted(ToBePopped, reverse=True):
            MasterList[indexMaster].pop(popItem)
        creation_string = "im createissue --type=\'Non Conforming Tracking - NCT\'" \
                          " --field=\'FA_Product Category=ECU\' " + " ".join(MasterList[indexMaster])
        NCT_number = [nct for nct in check_output(creation_string).split() if nct.isdigit()]
        print(NCT_number)
        NCT_OutputList.insert(0, "".join(NCT_number))

#Save Excel File TODO
Current_Date = datetime.datetime.now()
Tentative_title = "NCT_Created"\
                  + str(Current_Date.month)\
                  + "." + str(Current_Date.day)\
                  + "." + str(Current_Date.year)\
                  + ".xlsx"
print(Tentative_title)
print(NCT_OutputList)













